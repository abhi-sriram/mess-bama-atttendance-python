import openpyxl

path = "data.xlsx" # Raw data sheet file path

wb_obj = openpyxl.load_workbook(path)

bm_data_obj = openpyxl.load_workbook("bama_data.xlsx") # Transformed data written excel file 

sheet_obj = wb_obj.active

# cell_obj = sheet_obj['A2':'B434']


dates = {}

for value in sheet_obj.iter_rows(values_only=True):
    if(value[1]==None):
        continue
    if(str(value[1]).split(' ')[0] in dates):
        dates[str(value[1]).split(' ')[0]][0].append(
            str(value[0]).strip())
        dates[str(value[1]).split(' ')[0]][1].append(
            str(value[1]).strip())
    else:
        dates[str(value[1]).split(' ')[0]] = [
            [str(value[0]).strip()], [str(value[1]).strip()]]
# print(dates)

# for cell1, cell2 in cell_obj:
#     if(str(cell2.value).split(' ')[0] in dates):
#         dates[str(cell2.value).split(' ')[0]][0].append(str(cell1.value).strip())
#         dates[str(cell2.value).split(' ')[0]][1].append(str(cell2.value).strip())
#     else:
#         dates[str(cell2.value).split(' ')[0]] = [[str(cell1.value).strip()], [str(cell2.value).strip()]]


'''
06-11 : bf
12-15 : ln
16-18 : sn
19-23 : dn
'''
for date, data in dates.items():
    if(str(date) not in bm_data_obj.sheetnames):
        bm_data_obj.create_sheet(str(date))
    bm_sheet_obj = bm_data_obj[str(date)]
    bm_sheet_obj.cell(row=1, column=1).value = "Student ID"
    bm_sheet_obj.cell(row=1, column=2).value = "Breakfast"
    bm_sheet_obj.cell(row=1, column=3).value = "Lunch"
    bm_sheet_obj.cell(row=1, column=4).value = "Snack"
    bm_sheet_obj.cell(row=1, column=5).value = "Dinner"
    bm_sheet_obj.cell(row=1, column=6).value = "Total"

    bama_data = {
        # 'b172197': {
        #     'bf': 0,
        #     'ln': 0,
        #     'sn': 0,
        #     'dn': 0,
        #     'total': 0
        # }
    }
    ids = data[0]
    timestamps = data[1]
    bf = 0
    ln = 0
    sn = 0
    dn = 0
    for id, timestamp in zip(ids, timestamps):
        s_ts = timestamp.split(' ')[1].split(':')

        if(id in bama_data):
            if(int(s_ts[0]) >= 6 and int(s_ts[0]) <= 11 and bama_data[id]['bf'] == 0):
                bama_data[id]['bf'] = 1
                bama_data[id]['total'] += 1
                bf += 1
            elif(int(s_ts[0]) >= 12 and int(s_ts[0]) <= 15 and bama_data[id]['ln'] == 0):
                bama_data[id]['ln'] = 1
                bama_data[id]['total'] += 1
                ln += 1
            elif(int(s_ts[0]) >= 16 and int(s_ts[0]) <= 18 and bama_data[id]['sn'] == 0):
                bama_data[id]['sn'] = 1
                bama_data[id]['total'] += 1
                sn += 1
            elif(int(s_ts[0]) >= 19 and int(s_ts[0]) <= 23 and bama_data[id]['dn'] == 0):
                bama_data[id]['dn'] = 1
                bama_data[id]['total'] += 1
                dn += 1
        else:
            bama_data[id] = {
                'bf': 0,
                'ln': 0,
                'sn': 0,
                'dn': 0,
                'total': 0
            }
            if(int(s_ts[0]) >= 6 and int(s_ts[0]) <= 11):
                bama_data[id]['bf'] += 1
                bama_data[id]['total'] += 1
                bf += 1

            elif(int(s_ts[0]) >= 12 and int(s_ts[0]) <= 15):
                bama_data[id]['ln'] += 1
                bama_data[id]['total'] += 1
                ln += 1
            elif(int(s_ts[0]) >= 16 and int(s_ts[0]) <= 18):
                bama_data[id]['sn'] += 1
                bama_data[id]['total'] += 1
                sn += 1
            elif(int(s_ts[0]) >= 19 and int(s_ts[0]) <= 23):
                bama_data[id]['dn'] += 1
                bama_data[id]['total'] += 1
                dn += 1
    keys = list(bama_data.keys())
    values = list(bama_data.values())

    for i in range(len(keys)):
        bm_sheet_obj.cell(row=i+2, column=1).value = keys[i]
        bm_sheet_obj.cell(row=i+2, column=2).value = values[i]['bf']
        bm_sheet_obj.cell(row=i+2, column=3).value = values[i]['ln']
        bm_sheet_obj.cell(row=i+2, column=4).value = values[i]['sn']
        bm_sheet_obj.cell(row=i+2, column=5).value = values[i]['dn']
        bm_sheet_obj.cell(row=i+2, column=6).value = values[i]['total']
    bm_sheet_obj.cell(row=len(keys)+2, column=1).value = "Total"
    bm_sheet_obj.cell(row=len(keys)+2, column=2).value = bf
    bm_sheet_obj.cell(row=len(keys)+2, column=3).value = ln
    bm_sheet_obj.cell(row=len(keys)+2, column=4).value = sn
    bm_sheet_obj.cell(row=len(keys)+2, column=5).value = dn
    bm_data_obj.save("bama_data.xlsx")

# for date,data in bama_data.items():
#     print(date)
#     for i in range(len(data.values())):
#         print(data.values())

# for i in range(len(bama_data.values())):
    # print(bama_data.keys())
    # print(bama_data.values())
